#!/usr/bin/env python3
"""
Fielder Twitter Harassment Categorization Script
=================================================
Categorizes tweets directed at/about SF Supervisor Jackie Fielder
into 44 abuse/harassment categories and scores hostility 1-10.

Usage:
    python categorize_tweets.py input.csv output.xlsx

Input CSV must have at minimum a 'Reply Text' column and a 'Flag' column.
- Rows with Flag = "Yes" or "yes" will be categorized and scored.
- Rows with Flag = "No" or "no" will be skipped.
- Rows with any other flag (H, blank, Maybe, etc.) will be tested against
  the regex patterns; if any match, they'll be flagged "Maybe".

Output is an Excel file with categories and hostility scores filled in.
"""

import re
import sys
import pandas as pd
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# TAXONOMY: 44 categories across 4 groups
# ============================================================
#
# GENDERED / GENDER-BASED (11):
#   Gendered slur, Gendered slur + political, Hysteria,
#   Gendered nickname, Sexualization, Sexual harassment,
#   Appearance-based, Gendered dismissal, Gender-wide stereotype,
#   Infantilization, Comparison to other women
#
# IDENTITY-BASED (6):
#   Homophobia, Transphobia, Identity denial,
#   Identity delegitimization, Weaponizing sexual violence,
#   Racist comment
#
# INTELLIGENCE / COMPETENCE (3):
#   Attacking intelligence, Competence attack, Unqualified
#
# NON-GENDERED HOSTILITY (24):
#   Aggressive insult, Direct profanity, Ableist slur,
#   Dehumanizing, Hatred, Calling her crazy/insane,
#   Ideology as mental illness, Mocking mental illness,
#   Armchair diagnosis, Celebrating hospitalization,
#   Faking illness accusation, Death/harm wish,
#   Deportation threat, Aggressive dismissal,
#   Exaggerated criticism, Condescension,
#   Grifting accusation, Corruption accusation,
#   Criminal accusation, Removal demand,
#   Antisemitism allegation, Aggressive political slur,
#   Lying accusation, Blaming for Trump
# ============================================================

# Categories that count as gendered/identity-based for analysis
GENDERED_CATS = {
    'Gendered slur', 'Gendered slur + political', 'Hysteria',
    'Gendered nickname', 'Sexualization', 'Sexual harassment',
    'Appearance-based', 'Gendered dismissal', 'Gender-wide stereotype',
    'Infantilization', 'Comparison to other women',
    'Homophobia', 'Transphobia', 'Weaponizing sexual violence',
    'Identity denial', 'Identity delegitimization',
}

# Normalize variant category names (from manual tagging)
NORMALIZE = {
    'Exaggerated critcism': 'Exaggerated criticism',
    'Exaggerated criticsm': 'Exaggerated criticism',
    'Extreme criticism': 'Exaggerated criticism',
    'Condescending': 'Condescension',
    'Condecension': 'Condescension',
    'Infantilizing': 'Infantilization',
    'Comptence attack': 'Competence attack',
    'Grifting allegation': 'Grifting accusation',
    'Grifting accusations': 'Grifting accusation',
    'Corruption allegation': 'Corruption accusation',
    'Dehumanization': 'Dehumanizing',
    'Gender-based stereotype': 'Gender-wide stereotype',
    'Insult': 'Aggressive insult',
    'Faking illness': 'Faking illness accusation',
    'Calling socialism a mental illness': 'Ideology as mental illness',
    'Homophobic slur': 'Homophobia',
}


def categorize(text):
    """
    Classify a tweet into one or more of 44 harassment/abuse categories.
    Returns a semicolon-separated string of category names, sorted alphabetically.
    """
    tl = text.lower()
    cats = set()

    # ==============================================================
    # GENDERED / GENDER-BASED
    # ==============================================================

    # --- Gendered slur ---
    # Slurs targeting her gender: cunt, bitch, whore, slut, skank, hag,
    # twat, wench, hoe (including misspellings like cvnt, biatch, biotch)
    if re.search(
        r'\bcunt\b|\bcvnt\b|\bcuntbitch\b|\bc\*nt\b'
        r'|\bbitch\b|\bbiatch\b|\bbiotch\b|\bbitches\b'
        r'|\bwhore\b|\bslut\b|\bskank\b|\bhag\b|\btwat\b'
        r'|\bho\s*bag\b|\bwench\b|\bhoe\b|\bputa\b|\bsalesputa\b',
        tl
    ):
        cats.add('Gendered slur')

    # --- Slut/skank → also Sexualization ---
    # These slurs are inherently sexualizing
    if re.search(r'\bslut\b|\bskank\b', tl):
        cats.add('Sexualization')

    # --- Gendered slur + political ---
    # A gendered slur fused with a political label in the same attack
    if re.search(
        r'(communist|socialist|commie|nazi|fascist|liberal|leftist|marxist)'
        r'.{0,30}(cunt|cvnt|bitch|whore|slut|skank)'
        r'|(cunt|cvnt|bitch|whore|slut|skank)'
        r'.{0,30}(communist|socialist|commie|nazi|fascist|liberal|leftist|marxist)',
        tl
    ):
        cats.add('Gendered slur + political')

    # --- Hysteria ---
    # Language rooted in the "hysterical woman" trope — framing her
    # emotions or agency as pathological
    if re.search(
        r'\bhysterical\b|emotionally\s*driven|froth\s*at\s*the\s*mouth'
        r'|so\s*sensitive|total\s*inability\s*to\s*laugh'
        r'|cry\s*herself\s*out\s*of|flustered',
        tl
    ):
        cats.add('Hysteria')

    # --- Gendered nickname ---
    # Diminutive or mocking gendered names that reduce her to a
    # gendered caricature
    if re.search(
        r'#?hitlerbarbie|\bbanshee\b|princess\s*of\s*death'
        r'|\bsocialista\b|#?girl\s*boss|jackie\s*girl'
        r'|the\s*lady\b|says\s*the\s*lady|miss\s*keep|jackoff'
        r'|crazy\s*(chick|girl|lady|woman)|pap\s*smear',
        tl
    ):
        cats.add('Gendered nickname')
    # "estupida" = gendered Spanish insult (feminine -a ending)
    if re.search(r'\bestupida\b', tl):
        cats.add('Gendered nickname')

    # --- Sexualization ---
    # Comments about her attractiveness or desirability in a political context
    if re.search(
        r'she.s\s*(hot|cute|attractive|pretty)|hotttt'
        r'|hot\s*(chick|dumb)|eye\s*candy|wild\s*in\s*the\s*sack'
        r'|easy\s*on\s*the\s*eyes|🥵|lipstick\s*lesbian|tease\b|cumunist',
        tl
    ):
        cats.add('Sexualization')

    # --- Sexual harassment ---
    # Predatory or soliciting messages directed at her
    if re.search(
        r'(dm\s*me|send\s*me).{0,30}(bikini|lingerie|photo|pic)'
        r'|sugar\s*daddy\s*hmu|hot\s*teacher.*hmu'
        r'|put\s*her\s*in.*bikini|pack\s*lightly.*bikini'
        r'|essentials.*jackietown',
        tl
    ):
        cats.add('Sexual harassment')

    # --- Appearance-based ---
    # Attacks on physical appearance, body, or hygiene
    if re.search(
        r'\bugly\b|skeletor|five\s*head|forehead.*(large|alarming)'
        r'|got\s*ugly|horse\s*gif|derp\b|tattoo.*(crazy|cutter)'
        r'|knife\s*tattoo|pouty|annoying.*face|menopause'
        r'|vagina\s*removed|that\s*tattoo\b',
        tl
    ):
        cats.add('Appearance-based')
    # Smell/hygiene — but NOT figurative "I smell BS"
    if re.search(r'(smell|stink|smelly|stank|shower|hygiene|arm\s*pit|armpits)', tl):
        if re.search(r'(you|she|her|screen|through|jackie|fielder|communist\s*hippie)', tl):
            if not re.search(r'i\s*smell\s*(bs|bullshit|a\s*rat|something|trouble)', tl):
                cats.add('Appearance-based')
    if re.search(r'quiet\s*piggy|\boink\b|piggy', tl):
        cats.add('Appearance-based')
    if re.search(r'septum\s*pierc|turquoise\s*hair|botox|fillers|nose\s*job', tl):
        cats.add('Appearance-based')
    if re.search(r'(look|looks)\s*(like|terrible|awful|high)', tl):
        if re.search(r'(she|jackie|fielder|you)', tl):
            cats.add('Appearance-based')

    # --- Gendered dismissal ---
    # Using her gender, reproductive status, or womanhood as a disqualifier
    if re.search(
        r'childless|baby\s*mama|female\s*card|cat\s*lady|big\s*girl\s*job'
        r'|excuse\s*me\s*sir|empty\s*egg\s*carton|abolish\s*the\s*19th',
        tl
    ):
        cats.add('Gendered dismissal')

    # --- Gender-wide stereotype ---
    # Generalizing from her to all women or "Democrat women"
    if re.search(
        r'women\s*(don.t|do\s*not|shouldn.t)\s*belong'
        r'|democrat\s*women\b.*(mental|crazy|ill)'
        r'|liberal\s*women\b.*(mental|crazy|ill)'
        r'|typical\s*female|post.trump\s*women\s*descend'
        r'|these\s*kind\s*of\s*women|female.*socialist.*parody'
        r'|woman.s\s*fuck.up.*cry\s*herself'
        r'|midwit\s*woman',
        tl
    ):
        cats.add('Gender-wide stereotype')

    # --- Infantilization ---
    # Framing an elected official in her 30s as a child or child-like
    if re.search(
        r'temper\s*tantrum|(?:she|jackie)\s*is\s*a\s*child'
        r'|what\s*an?\s*overgrown\s*child|poor\s*baby|overgrown\s*child'
        r'|adulting|13.year.old|little\s*girl|brainwashed\s*child'
        r'|music\s*video\s*extra|cosplay.*politi|college\s*idiot'
        r'|jackie\s*girl|4\s*yo\s*has\s*better|child.s\s*(interpretation|thinking)'
        r'|the\s*child\s*(doesn|don)|theater\s*kid|you.re\s*a.*brat',
        tl
    ):
        cats.add('Infantilization')
    if re.search(r'(politics|being\s*a\s*politician)\s*(is\s*)?(really\s*)?(not\s*for\s*everyone|not\s*for\s*her)', tl):
        cats.add('Infantilization')
    if re.search(r'can.t\s*take\s*you\s*seriously.*as\s*an\s*adult', tl):
        cats.add('Infantilization')
    if re.search(r'economically\s*illiterate\s*and\s*childish|just.*childish', tl):
        cats.add('Infantilization')
    # Mocking imitation of her voice = Infantilization + Condescension
    if re.search(r'hehehe|look\s*how\s*smart\s*i\s*am', tl):
        cats.add('Infantilization')
        cats.add('Condescension')

    # --- Comparison to other women ---
    # Comparing or ranking her against other women politicians
    if re.search(
        r'(just\s*like|same\s*as|worse\s*than|crazier\s*than|indistinguishable.*from)'
        r'\s*(hillary|aoc|pelosi|warren|ronen|breed|harris|lina\s*hidalgo|lorena)',
        tl
    ):
        cats.add('Comparison to other women')
    if re.search(r'(hotttt?|hot|cute|attractive).*(aoc|pelosi|boebert|harris)', tl):
        cats.add('Comparison to other women')
    if re.search(r'who.s\s*this\s*bitch.*aoc|aoc.*who.s\s*this\s*bitch', tl):
        cats.add('Comparison to other women')
    if re.search(r'ketamine\s*therapy\s*with\s*@?aoc|@?aoc\s*is\s*dumb', tl):
        cats.add('Comparison to other women')

    # ==============================================================
    # IDENTITY-BASED
    # ==============================================================

    # --- Homophobia ---
    # Slurs + broader homophobic hostility
    if re.search(r'\bdyke\b|\blesbo\b|\bbutch\b|\bso\s*gay\b|Gay!\b', tl):
        cats.add('Homophobia')
    if re.search(r'(another\s*sf\s*gay|sf\s*gay\s*has)', tl):
        cats.add('Homophobia')
    if re.search(
        r'sex\s*change\s*operation|domestic.*partners.*\?'
        r'|rainbow\s*flag.*parody|lesbian.*looking|lesbian.*box'
        r'|false\s*flag\s*op.*profile',
        tl
    ):
        cats.add('Homophobia')

    # --- Transphobia ---
    # Misgendering, implying male anatomy, mocking trans healthcare
    if re.search(
        r'testicular\s*cancer|excuse\s*me\s*sir|calling\s*her\s*a\s*man'
        r'|sex\s*change|he\s*\/\s*him',
        tl
    ):
        cats.add('Transphobia')

    # --- Identity denial ---
    # Denying or attacking her claimed identities
    if re.search(
        r'fake\s*indigenous|pretendian|fake\s*indian|fake\s*native'
        r'|fake\s*queer|fake\s*lesbian|carpetbag|white\s*divorc'
        r'|te\s*falta\s*barrio|stanford\s*fresa',
        tl
    ):
        cats.add('Identity denial')

    # --- Identity delegitimization ---
    # Framing her identities as box-checking
    if re.search(
        r'checks?\s*all\s*the\s*boxes|ticked.*boxes|oppression\s*olympics'
        r'|virtue\s*signal.*identit|identity\s*politics\s*propaganda'
        r'|female.*socialist.*rainbow.*parody',
        tl
    ):
        cats.add('Identity delegitimization')

    # --- Weaponizing sexual violence ---
    # Holding her responsible for men's violence or invoking it
    if re.search(
        r'rapist\s*(enabler|supporter|friend)'
        r'|jacobo.{0,20}rapist|rapist.{0,20}jacobo'
        r'|simp\s*for\s*rapist|credibly\s*accused\s*sexual\s*assault'
        r'|financially\s*raping',
        tl
    ):
        cats.add('Weaponizing sexual violence')
    if re.search(r'women\s*and\s*children\s*raped\s*and\s*murdered', tl):
        if re.search(r'(ice|deport|immigra)', tl):
            cats.add('Weaponizing sexual violence')
            cats.add('Racist comment')

    # --- Racist comment ---
    # Racial slurs, nativist hostility, pro-deportation/ICE rhetoric
    if re.search(
        r'\bhondos\b|honduran\s*drug\s*dealer|kate\s*steinle'
        r'|your\s*people|white\s*people\s*are\s*getting\s*tired'
        r'|seasoning\s*and\s*flavor|illegal\s*alien.*murderer'
        r'|queer.*dei.*nonsense|keep\s*the\s*mission\s*mexican'
        r'|hispanic\s*looking|white\s*name|mexican\s*flag\s*in\s*bio'
        r'|monsters\s*in\s*the\s*mission',
        tl
    ):
        cats.add('Racist comment')
    # General pro-deportation/pro-ICE rhetoric (not personal threats)
    if re.search(
        r'(deport|remove|kick\s*out)\s*(them\s*all|all\s*(of\s*them|illegal|undocumented))'
        r'|more\s*ice|go\s*ice|nah.*i\s*like\s*borders'
        r'|nah.*ice\s*is\s*welcome|ice\s*is\s*not\s*going\s*anywhere'
        r'|ice\s*is\s*welcome|(abolish|abolishing)\s*ice'
        r'|ice\s*blah|go\s*seahawks.*ice|go\s*i\.?c\.?e',
        tl
    ):
        cats.add('Racist comment')
    if re.search(r'(illegal|illegals|undocumented).*(deport|remove|out|go\s*home|go\s*back)', tl):
        cats.add('Racist comment')
    if re.search(r'(deport|remove|illegal|illegals).*(immigrant|alien|border\s*crosser)', tl):
        cats.add('Racist comment')
    if re.search(r'millions\s*are\s*being\s*deported', tl):
        cats.add('Racist comment')
    if re.search(r'(need|needs|desperate)\s*(of\s*)?ice\b', tl):
        cats.add('Racist comment')
    if re.search(r'maid\s*can\s*stay', tl):
        cats.add('Racist comment')
    if re.search(r'(pro\s*homeless|protect.*illegal|protecting.*criminal|protect.*drug\s*dealer)', tl):
        if re.search(r'(ice|deport|immigra)', tl):
            cats.add('Racist comment')

    # ==============================================================
    # INTELLIGENCE / COMPETENCE
    # ==============================================================

    # --- Attacking intelligence ---
    # Direct insults (stupid, dumb, idiot, moron, 🤡) AND implied
    # attacks (take an econ class, you don't understand)
    if re.search(
        r'\b(stupid|dumb|dumbest|dopey|dimb|brain\s*dead|idiot|moron'
        r'|imbecile|halfwit|dimwit|nitwit|dunce|dipshit|dipstick'
        r'|dumbass|midwit)\b',
        tl
    ):
        cats.add('Attacking intelligence')
    if re.search(r'economically\s*(illiterate|ill?iterate)|financially\s*illiterate', tl):
        cats.add('Attacking intelligence')
    if re.search(r'🤡', tl):
        cats.add('Attacking intelligence')
    if re.search(r'(take|try)\s*(a|some)\s*(basic\s*)?(economics?|econ|math)\s*(class|101|course)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'(you\s*(don.t|have\s*no|lack)\s*(understand|idea|clue|concept|knowledge))', tl):
        cats.add('Attacking intelligence')
    if re.search(r'do\s*you\s*(even\s*)?(understand|know|comprehend)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'(read|learn|study)\s*(some\s*)?(economics|econ|the\s*constitution|a\s*book|history)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'tell\s*me\s*you.*(don.t|without)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'(not\s*(smart|bright|intelligent)\s*enough|below\s*\d+.*iq|pea.brain|low\s*iq|zero.*iq)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'showing\s*your\s*ignorance|\bignorant\b', tl):
        cats.add('Attacking intelligence')
    if re.search(r'\bclueless\b', tl):
        cats.add('Attacking intelligence')
    if re.search(r'not\s*a\s*serious\s*(person|human)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'\bestupida\b', tl):
        cats.add('Attacking intelligence')
    if re.search(r'first\s*grade\s*econ', tl):
        cats.add('Attacking intelligence')
        cats.add('Condescension')
    if re.search(r'fuck\s*off\s*with\s*thinking', tl):
        cats.add('Attacking intelligence')
    if re.search(r'@ifindretards', tl):
        cats.add('Attacking intelligence')
    if re.search(r'(what\s*a|what\s*an?)\s*maroon', tl):
        cats.add('Attacking intelligence')
    if re.search(r'appalling.*knowledge|knowledge.*appalling|misunderstand', tl):
        cats.add('Attacking intelligence')
    if re.search(r'not\s*a\s*deep\s*thinker', tl):
        cats.add('Attacking intelligence')
    if re.search(r'(you.re|she.s)\s*(surprisingly\s*)?(dumb|stupid|ignorant)', tl):
        cats.add('Attacking intelligence')
    if re.search(r'idiot\s*sandwich', tl):
        cats.add('Attacking intelligence')

    # --- Ableist slur ---
    # ALWAYS co-occurs with Attacking intelligence
    if re.search(r'\bretard(ed|s|io|ios)?\b|@ifindretards|re\.\s*tard|full\s*retard', tl):
        cats.add('Ableist slur')
        cats.add('Attacking intelligence')

    # --- Competence attack ---
    # Attacking her ability to do her job or claiming she does nothing
    if re.search(r'(incompetent|useless|feckless|inept|do.nothing|negligent|neglect)', tl):
        cats.add('Competence attack')
    if re.search(r'(does?|did|has)\s*(nothing|jack\s*shit|fuck\s*all|zero|nada|squat|absolutely\s*nothing|not\s*a\s*damn)', tl):
        cats.add('Competence attack')
    if re.search(
        r'(do\s*your\s*(damn|fucking)?\s*job|do\s*something|do\s*better'
        r'|zero\s*(accomplish|achiev)|what\s*does?\s*she\s*do\s*all\s*day)',
        tl
    ):
        cats.add('Competence attack')
    if re.search(r'(lazy|absent|missing|awol|no\s*show)\b', tl):
        if re.search(r'(super|politi|supe|leader|official|grift|jackie|fielder|you|she)', tl):
            cats.add('Competence attack')
    if re.search(r'scatterbrain|airhead|bimbo|ditz', tl):
        cats.add('Competence attack')
    if re.search(r'(can.t|couldn.t|doesn.t)\s*(run|manage|handle|organize|do\s*(math|her\s*job|anything))', tl):
        cats.add('Competence attack')
    if re.search(r'haven.t\s*done\s*(shit|anything|a\s*thing)|didn.t\s*do\s*(shit|anything)', tl):
        cats.add('Competence attack')

    # --- Unqualified ---
    # Not fit or qualified to hold office
    if re.search(r'(not|never|zero|no)\s*(qualified|fit|competent|capable|suited)\s*(for|to)', tl):
        cats.add('Unqualified')
    if re.search(r'shouldn.t\s*be\s*(anywhere\s*)?near\s*government', tl):
        cats.add('Unqualified')
    if re.search(r'big\s*girl\s*job', tl):
        cats.add('Unqualified')
    if re.search(r'never\s*(fit|qualified)\s*to\s*govern', tl):
        cats.add('Unqualified')
    if re.search(r'get\s*a\s*real\s*job', tl):
        cats.add('Unqualified')
    if re.search(r'should\s*not\s*be\s*allowed\s*to\s*vote', tl):
        cats.add('Unqualified')

    # ==============================================================
    # NON-GENDERED HOSTILITY
    # ==============================================================

    # --- Aggressive insult ---
    # Non-intelligence personal insults
    if re.search(r'\b(loser|coward|jackass|fool|buffoon|clown)\b', tl):
        cats.add('Aggressive insult')
    if re.search(r'(what\s*a|such\s*a|total|complete|absolute|fucking)\s*(disaster|embarrassment|joke|failure|disgrace)', tl):
        cats.add('Aggressive insult')
    if re.search(r'\bestupida\b|\bpendeja\b', tl):
        cats.add('Aggressive insult')
    if re.search(r'(terrible|horrible|awful|miserable)\s*person', tl):
        cats.add('Aggressive insult')
    if re.search(r'entitled\b', tl) and re.search(r'(jackie|fielder|she|you|bitch)', tl):
        cats.add('Aggressive insult')
    if re.search(r'job\s*killer', tl):
        cats.add('Aggressive insult')
    if re.search(r'couldn.t\s*happen\s*to\s*a\s*wors', tl):
        cats.add('Aggressive insult')

    # --- Direct profanity ---
    # Fuck, shit, and variants (gendered slurs categorized separately)
    if re.search(r'\bfuck\b|\bfucking\b|\bfuckin\b|\bfking\b|\bf\*ck|fekk|fck\b|f.ck\b', tl):
        cats.add('Direct profanity')
    if re.search(r'\bstfu\b|\bgtfo\b|\bgfy\b|🖕|🤬|fu\s+loser|\bfu\b$|\bFU\b', tl):
        cats.add('Direct profanity')
    if re.search(r'\bshit\b|\bshitty\b|\bshithead\b|\bshithole\b', tl):
        if re.search(r'(you|she|her|jackie|fielder|city|district|show)', tl):
            cats.add('Direct profanity')
    if re.search(r'wtf\b|azwipe', tl):
        cats.add('Direct profanity')
    if re.search(r'\.\.\.and\s*the\s*horse\s*you\s*rode', tl):
        cats.add('Direct profanity')

    # --- Aggressive dismissal ---
    # Commanding her to be silent or go away
    # "fuck off" / "GTFO" = always aggressive dismissal
    if re.search(r'fuck\s*off|gtfo|just\s*stay\s*out|bug\s*off|gfy', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'shut\s*(the\s*fuck\s*)?up|zip\s*it|pipe\s*down|shuddup|shup\s*up|stfu', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'nobody\s*(likes|cares\s*about|wants|asked|takes.*seriously)', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'no\s*one\s*(cares|wants|takes|likes)', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'sit\s*down\b|stop\s*wasting\s*oxygen', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'up\s*yours\b|quiet\s*demon', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'(you|she)\s*(are|is)\s*not\s*serious', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'quiet\s*piggy', tl):
        cats.add('Aggressive dismissal')
    if re.search(r'opinion\s*discarded', tl):
        cats.add('Aggressive dismissal')

    # --- Dehumanizing ---
    # Language that strips away her personhood
    # But NOT when "disgusting" refers to an incident/conditions
    if re.search(
        r'(you.re|she.s|you\s*are|she\s*is|jackie\s*is)\s*'
        r'(worthless|trash|garbage|a\s*disgrace|pathetic|disgusting'
        r'|a\s*waste|evil|vile|scum|despicable|a\s*bad\s*person'
        r'|a\s*cancer|a\s*stain)',
        tl
    ):
        cats.add('Dehumanizing')
    if re.search(r'worthless\s*(jackie|fielder|supervisor|politician|elected)', tl):
        cats.add('Dehumanizing')
    if re.search(r'(\btrash\b|\bgarbage\b|\bscum\b|\bghoul\b|\bparasite\b|\bleech\b|\bvermin\b)', tl):
        if re.search(r'(you|she|her|jackie|fielder|@jackiefielder|later\s*trash|elect\s*this)', tl):
            cats.add('Dehumanizing')
    if re.search(r'waste\s*of\s*(space|skin|oxygen|air|time|water|calcium)', tl):
        cats.add('Dehumanizing')
    if re.search(r'piece\s*of\s*(shit|crap|garbage)|barely\s*human|subhuman|sludge\s*monster', tl):
        cats.add('Dehumanizing')
    if re.search(r'(evil|cancer|tumor|plague|poison|stain)\b', tl):
        if re.search(r'(jackie|fielder|you|she|supervisor|your\s*bio)', tl):
            cats.add('Dehumanizing')
    if re.search(r'(you|she|jackie)\s*(suck|are\s*a\s*bad\s*person|make\s*me\s*sick)', tl):
        cats.add('Dehumanizing')
    if re.search(r'disgrace\s*to\s*(the|human)', tl):
        cats.add('Dehumanizing')
    if re.search(r'drooling', tl):
        cats.add('Dehumanizing')
    if re.search(r'quiet\s*demon', tl):
        cats.add('Dehumanizing')
    # "disgusting" about conditions/incidents = NOT dehumanizing
    if re.search(r'(disgusting|pathetic|worthless)\b', tl):
        if re.search(r'(you|she|her|jackie|fielder|@jackiefielder)', tl):
            if not re.search(
                r'(antisemit|anti.semit|chant|slur|moment\s*of|mask\s*off'
                r'|conditions|district|streets|sidewalk|magnet\s*for\s*misery'
                r'|3rd\s*world|enabling\s*death)',
                tl
            ):
                cats.add('Dehumanizing')
    if re.search(r'(extremism|extremist).*root.*evil', tl):
        cats.add('Dehumanizing')

    # --- Hatred ---
    # Direct expressions of personal hatred
    if re.search(r'(i\s*|we\s*)(hate|loathe|despise|detest|abhor)\s*(you|her|jackie|fielder)', tl):
        cats.add('Hatred')
    if re.search(r'h8.s\s*this\s*woman', tl):
        cats.add('Hatred')
    if re.search(r'(you|she|jackie|fielder)\s*(disgust|sicken)\s*(me|us)', tl):
        cats.add('Hatred')

    # --- Mental illness subcategories ---
    # Context checks to avoid false positives
    is_street = bool(re.search(
        r'(street|sidewalk|block|bart|muni|16th|24th|mission\s*district)'
        r'.{0,50}(addict|junkie|drug|homeless|mentally\s*ill)',
        tl
    ))
    is_dsa_debate = bool(re.search(
        r'blame.*mental.*homeless|homeless\s*woman|bystander|random.*mental',
        tl
    ))
    is_colloquial_crazy = bool(re.search(
        r'(it.s|that.s|this\s*is)\s*(crazy|insane)\s*(how|that|to)',
        tl
    ))

    # Mocking mental illness
    if re.search(
        r'(not\s*easy|hard)\s*being\s*(crazy|insane)|she\s*cray'
        r'|tds\b|trump\s*derangement|waymo\s*derangement'
        r'|meds\s*kick\s*in|gastrointestinal|psych\s*unit'
        r'|mental\s*ward|suicidal\s*compassion'
        r'|stick.*removed.*ass|mental\s*health\s*condition.*communist'
        r'|fielder\s*in\s*sf\s*general',
        tl
    ):
        cats.add('Mocking mental illness')
    if re.search(r'(lol|lmao|haha|😆|🤣|😂).{0,50}(mental|hospital|crisis|breakdown)', tl):
        cats.add('Mocking mental illness')
    if re.search(r'(mental|hospital|crisis|breakdown).{0,50}(lol|lmao|haha|😆|🤣|😂)', tl):
        cats.add('Mocking mental illness')
    if re.search(r'nature\s*is\s*healing', tl):
        cats.add('Mocking mental illness')
        cats.add('Celebrating hospitalization')

    # Ideology as mental illness
    if re.search(
        r'(socialist|leftist|liberal|dsa|democrat|progressive|communi)'
        r'.{0,30}(insane|crazy|mental|lunatic|nuts|deranged|unhinged'
        r'|psycho|loony|nutter|unstable|illness|disorder|breakdown)',
        tl
    ):
        cats.add('Ideology as mental illness')
    if re.search(
        r'(insane|crazy|mental|lunatic|nuts|deranged|unhinged|psycho'
        r'|loony|nutter|unstable)'
        r'.{0,30}(socialist|leftist|liberal|dsa|democrat|progressive|communi)',
        tl
    ):
        cats.add('Ideology as mental illness')

    # Calling her crazy/insane (personal, not colloquial)
    if not is_street and not is_dsa_debate and not is_colloquial_crazy:
        if re.search(
            r'(she|you|jackie|fielder)\s*(is|are|seems?|clearly)\s*'
            r'(insane|crazy|deranged|unhinged|psycho|a\s*lunatic|mental'
            r'|nuts|delusional|batshit|loony|bonkers|wacko|nutjob|nutter)',
            tl
        ):
            cats.add('Calling her crazy/insane')
        if re.search(r'mentally\s*(ill|unstable|unfit|unbalanced|disturbed)', tl):
            if not re.search(r'(people|folks|individuals|homeless|addicts|those)\s*(who\s*are\s*)?mentally', tl):
                cats.add('Calling her crazy/insane')
        if re.search(r'nuttier\s*than|crazier\s*than\s*a|lost\s*her\s*marbles', tl):
            cats.add('Calling her crazy/insane')
        if re.search(r'psycho\s*bitch|insane\s*bitch', tl):
            cats.add('Calling her crazy/insane')
        if re.search(r'came\s*to\s*her\s*senses', tl):
            cats.add('Calling her crazy/insane')
        if re.search(r'the\s*insanity\b', tl):
            cats.add('Calling her crazy/insane')

    # --- Celebrating hospitalization ---
    if re.search(
        r'god\s*is\s*good|lord\s*works\s*in\s*mysterious|good\s*riddance'
        r'|cause\s*for\s*celebration|miracle\s*in\s*disguise|bye\s*felecia'
        r'|too\s*bad.*libs',
        tl
    ):
        cats.add('Celebrating hospitalization')
    if re.search(r'(couldn.t happen|couldn.t\s*have\s*happened)\s*to\s*a\s*(worse|worst|better|nicer)', tl):
        cats.add('Celebrating hospitalization')
    if re.search(r'\bkarma\b', tl):
        if re.search(r'(jackie|fielder|hospital|mental|crisis|coming\s*her\s*way)', tl):
            cats.add('Celebrating hospitalization')

    # --- Faking illness accusation ---
    if re.search(
        r'(faking|fake|staged|act|stunt|bs|bullshit|playing|pretend'
        r'|pr\s*stunt|gimmick|cover|making.*up|convenient|scam.*sympathy)'
        r'.{0,40}(hospital|mental|illness|breakdown|crisis|sick|episode|recover|health)',
        tl
    ):
        cats.add('Faking illness accusation')
    if re.search(
        r'(hospital|mental|illness|breakdown|crisis|recover|health)'
        r'.{0,40}(faking|fake|staged|act|stunt|cover|convenient|scam)',
        tl
    ):
        cats.add('Faking illness accusation')
    if re.search(
        r'"hospitalized"|"recover|making\s*herself\s*into\s*a\s*victim'
        r'|using.*cover.*health|using.*excuse.*health|using.*sympathy'
        r'|she.s\s*clearly\s*dying',
        tl
    ):
        cats.add('Faking illness accusation')
    if re.search(r'woman.s\s*fuck.up.*cry\s*herself.*check\s*into\s*hospital', tl):
        cats.add('Faking illness accusation')

    # --- Armchair diagnosis ---
    if re.search(
        r'bipolar|bi-polar|schizophren|borderline\s*personality'
        r'|\bcutter\b|britney\s*spears|adachi|looks?\s*high'
        r'|narcissis|psychosis|drug\s*(test|rehab)|withdrawal'
        r'|illicit\s*drug|addiction',
        tl
    ):
        cats.add('Armchair diagnosis')

    # --- Death/harm wish ---
    if re.search(
        r'(should|hope|wish|pray|needs?\s*to|deserves?\s*to)\s*'
        r'(die|drop\s*dead|burn|rot|suffer|be\s*(killed|shot|hanged|executed))',
        tl
    ):
        cats.add('Death/harm wish')
    if re.search(r'kill\s*(yourself|herself)|euthanasia|put\s*(her|you)\s*down', tl):
        cats.add('Death/harm wish')
    if re.search(r'suffers?\s*miserably', tl):
        cats.add('Death/harm wish')
    if re.search(
        r'(send|release).*(gang|criminal|dealer).*your\s*(house|home)'
        r'|earthquake.*fix\s*everything',
        tl
    ):
        cats.add('Death/harm wish')
    if re.search(r'they\s*need\s*make\s*you\s*pay|we.re\s*going\s*to\s*make\s*them\s*pay', tl):
        cats.add('Death/harm wish')
    if re.search(r'I\s*hope\s*(jackie|fielder|she)\s*rots', tl):
        cats.add('Death/harm wish')
    if re.search(r'burn\s*in\s*hell', tl):
        cats.add('Death/harm wish')
    if re.search(r'(shoot|shooting)\s*(california\s*politic|politic)', tl):
        cats.add('Death/harm wish')
    # "Move to Cuba/Venezuela" = deportation threat AND death/harm wish
    if re.search(r'move\s*to\s*(cuba|venezuela)', tl):
        cats.add('Death/harm wish')
        cats.add('Deportation threat')

    # --- Deportation threat (personal only) ---
    if re.search(r'deport\s*(this|her|you|jackie|fielder)', tl):
        cats.add('Deportation threat')
    if re.search(r'(get\s*out|go\s*back)\s*(of\s*)?(my|our|the|your)\s*(country|state|city)', tl):
        cats.add('Deportation threat')
    if re.search(r'move\s*to\s*(minnesota|france)', tl):
        cats.add('Deportation threat')
    if re.search(r'(jackie|fielder)\s*please\s*leave|leave\s*america|gtfo\s*of\s*my\s*country', tl):
        cats.add('Deportation threat')

    # --- Exaggerated criticism ---
    if re.search(r'(worst|most\s*(useless|incompetent|pathetic))\s*(super|politi|supe|person|human|leader)', tl):
        cats.add('Exaggerated criticism')
    if re.search(r'blood\s*on\s*(your|her)\s*hands', tl):
        cats.add('Exaggerated criticism')
    if re.search(
        r'(destroying|ruining|wrecking|crushed|killed|killing)\s*'
        r'(sf|san\s*fran|the\s*(city|district|state|mission)|california)',
        tl
    ):
        cats.add('Exaggerated criticism')
    if re.search(r'(you.re|she.s|you\s*are|she\s*is)\s*(an?\s*)?(absolute\s*)?embarrassment', tl):
        cats.add('Exaggerated criticism')
    if re.search(r'(you|she|jackie|fielder)\s*(are|is)\s*(the|a|part\s*of\s*the)\s*problem', tl):
        cats.add('Exaggerated criticism')
    if re.search(
        r'(great\s*evil|evil\s*(crime|loving|socialist)|death\s*spiral'
        r'|dumpster\s*fire|shithole|zombie\s*apocalypse|3rd\s*world'
        r'|sphincter|laughing\s*stock|committed\s*great\s*evils)',
        tl
    ):
        cats.add('Exaggerated criticism')
    if re.search(r'worst\s*district|district\s*is\s*(a\s*)?(disaster|mess|rotten|filthy)', tl):
        cats.add('Exaggerated criticism')

    # --- Condescension ---
    if re.search(
        r'(sweetie|honey|dear|hun|bless\s*(your|her)\s*heart'
        r'|oh\s*gosh|i\s*feel\s*sorry\s*for\s*you|sweet\s*pea|sweat\s*pea)',
        tl
    ):
        cats.add('Condescension')
    if re.search(r'welcome\s*to\s*socialism|how.s\s*that\s*.*going|seek\s*help', tl):
        cats.add('Condescension')
    if re.search(r'oh\s*no!.*retarded\s*:\(', tl):
        cats.add('Condescension')
    if re.search(r'(parody|satirical|satyrical)\s*account\s*right', tl):
        cats.add('Condescension')

    # --- Grifting accusation ---
    if re.search(
        r'\bgrift(er|ing|s)?\b|showboat(ing)?'
        r'|performative\s*(nonsense|bs|bullshit)'
        r'|con\s*(artist|woman|job)|fail\s*up|champagne\s*socialist',
        tl
    ):
        cats.add('Grifting accusation')
    if re.search(r'perm\s*disability.*activism|disability.*free.*up|full\s*time\s*activism', tl):
        cats.add('Grifting accusation')
    if re.search(r'(public\s*sector\s*)?leech\s*wants\s*to\s*steal', tl):
        cats.add('Grifting accusation')
    if re.search(r'(virtue\s*signal|signaling)', tl) and not re.search(r'identity', tl):
        cats.add('Grifting accusation')
    if re.search(r'scam\s*to\s*get\s*sympathy', tl):
        cats.add('Grifting accusation')
    if re.search(r'(moral\s*blackmail|guilt.*disagree)', tl):
        cats.add('Grifting accusation')

    # --- Corruption accusation ---
    if re.search(
        r'(she|you|jackie|fielder).{0,30}'
        r'(corrupt|crooked|criminal|broke\s*the\s*law|lawbreak'
        r'|illegal\s*leak|violated|crook|thief|steal|stole|fraud)',
        tl
    ):
        cats.add('Corruption accusation')
    if re.search(
        r'(corrupt|crooked|criminal|crook|fraud|thief)'
        r'.{0,30}(she|her|jackie|fielder|politician|supervisor)',
        tl
    ):
        cats.add('Corruption accusation')
    if re.search(r'(steal|stole|stealing|theft|embezzle)\s*(money|funds|taxpayer)', tl):
        cats.add('Corruption accusation')
    if re.search(r'corruption\s*in\s*(this\s*state|politics|california|sf)', tl):
        cats.add('Corruption accusation')
    if re.search(r'crook\b|crooked\b', tl):
        cats.add('Corruption accusation')

    # --- Criminal accusation ---
    if re.search(
        r'(she|jackie|fielder|you)\s*(should|needs?\s*to|must|belongs?|deserves?)\s*(be\s*)?'
        r'(in\s*jail|arrested|locked\s*up|prosecuted|behind\s*bars|in\s*prison|indicted)',
        tl
    ):
        cats.add('Criminal accusation')
    if re.search(r'(arrest|jail|prosecute|indict|lock.*up)\s*(her|jackie|fielder)', tl):
        cats.add('Criminal accusation')
    if re.search(r'broke\s*the\s*law|breaking\s*the\s*law|illegally\s*leak|against\s*the\s*law|under\s*investigation', tl):
        cats.add('Criminal accusation')
    if re.search(r'communist\s*control\s*act', tl):
        cats.add('Criminal accusation')

    # --- Removal demand ---
    if re.search(r'(fire|recall|remove|oust|dump|boot|vote\s*out|get\s*rid\s*of|kick\s*out)\s*(her|jackie|fielder|this|all)', tl):
        cats.add('Removal demand')
    if re.search(
        r'(she|jackie|fielder)\s*(needs?\s*to|should|must|has\s*to)\s*'
        r'(resign|step\s*down|be\s*(removed|fired|recalled|replaced)|go|leave|quit)',
        tl
    ):
        cats.add('Removal demand')
    if re.search(
        r'#recall\s*jackie|recall\s*jackie|#recalljackiefielder'
        r'|condemn\s*or\s*resign|you\s*will\s*not\s*be\s*reelected'
        r'|vote\s*you\s*out|resign\s*in\s*disgrace',
        tl
    ):
        cats.add('Removal demand')
    if re.search(r'(jackie|fielder)\s*out\s*of\s*office', tl):
        cats.add('Removal demand')
    if re.search(r'fire\s*the\s*(fekking|fucking)\s*(commun|commie)', tl):
        cats.add('Removal demand')
    if re.search(r'please\s*resign|you.re\s*gone', tl):
        cats.add('Removal demand')
    if re.search(r'dump\s*jackie|dump\s*fielder', tl):
        cats.add('Removal demand')

    # --- Antisemitism allegation ---
    if re.search(
        r'(you|she|jackie|fielder|dsa|your\s*(org|party|group)|@dsa_sf|@demsocialists)'
        r'.{0,60}(nazi|antisemit|anti.semit|jew.hat|hate.*jews|anti.jewish)',
        tl
    ):
        cats.add('Antisemitism allegation')
    if re.search(
        r'(nazi|antisemit|anti.semit|jew.hat|anti.jewish)'
        r'.{0,60}(you|she|jackie|fielder|dsa|@dsa_sf)',
        tl
    ):
        cats.add('Antisemitism allegation')
    if re.search(r'ham\s*ass|#hamass|tax\s*the\s*jew', tl):
        cats.add('Antisemitism allegation')
    if re.search(r'(denounce|condemn).{0,30}(antisemit|anti.semit|chant|slur)', tl):
        if re.search(r'(jackie|fielder|dsa|@jackiefielder)', tl):
            cats.add('Antisemitism allegation')
    if re.search(r'disgusting.{0,30}(antisemit|anti.semit)', tl):
        cats.add('Antisemitism allegation')
    if re.search(r'disgusting.*slurs|anti\s*semitic\s*slurs|anti\s*semitic\s*shit', tl):
        if re.search(r'(dsa|jackie|fielder|@jackiefielder|@demsocialists)', tl):
            cats.add('Antisemitism allegation')

    # --- Aggressive political slur ---
    if re.search(r'(commie|communist|marxist)\s*(trash|scum|filth|rat|pig|bastard|hippie|piece)', tl):
        cats.add('Aggressive political slur')
    if re.search(r'(fucking|god\s*damn|fekking|fekk)\s*(commie|communist|marxist|socialist)', tl):
        cats.add('Aggressive political slur')
    if re.search(
        r'fire\s*the\s*(fekking|fucking)\s*(commun|commie)'
        r'|screw\s*you.*commie|🖕\s*commie|fuck\s*off\s*commie'
        r'|average\s*commie\s*take',
        tl
    ):
        cats.add('Aggressive political slur')
    if re.search(r'stinking\s*parasite.*fascist|fascist.*stinking\s*parasite', tl):
        cats.add('Aggressive political slur')

    # --- Lying accusation ---
    if re.search(r'(you.re|she.s|jackie.s?)\s*(a\s*)?(liar|lying|full\s*of\s*(shit|it)|dishonest|deceitful|gaslighting)', tl):
        cats.add('Lying accusation')
    if re.search(
        r'stop\s*lying|call\s*bullshit|bullshit\s*(on\s*this|lie)'
        r'|how\s*easily\s*you\s*lie|lying.*trying\s*to\s*use|lying\s*cunt',
        tl
    ):
        cats.add('Lying accusation')

    # --- Blaming for Trump ---
    if re.search(
        r'(why\s*we\s*(have|got)|reason.*we.*got|how\s*we\s*end\s*up\s*with'
        r'|no\s*wonder\s*we\s*(got|have)|thanks\s*to\s*you.*|de.facto\s*supported)\s*(trump|maga)',
        tl
    ):
        cats.add('Blaming for Trump')

    # ==============================================================
    # FALLBACK: if no categories matched, try broader patterns
    # ==============================================================
    if not cats:
        if re.search(r'(fuck|shit|damn|ass|hell)\b', tl):
            if re.search(r'(you|she|her|jackie|fielder)', tl):
                cats.add('Direct profanity')
        if re.search(r'(terrible|awful|horrible|worst|bad)\s*(person|human|supervisor|politician)', tl):
            cats.add('Exaggerated criticism')
        if re.search(r'(commie|communist|marxist|bolshevik)\b', tl):
            cats.add('Aggressive political slur')
        if re.search(r'(corrupt|crooked|fraud|steal|criminal)', tl):
            cats.add('Corruption accusation')
        if re.search(r'(resign|recall|fire|remove|vote\s*out|get\s*rid)', tl):
            cats.add('Removal demand')
        if re.search(r'(antisemit|nazi|jew)', tl):
            if re.search(r'(dsa|jackie|fielder|socialist)', tl):
                cats.add('Antisemitism allegation')
        if re.search(r'(stupid|dumb|idiot|moron|fool|clown|ignorant|clueless|brain\s*dead)', tl):
            cats.add('Attacking intelligence')
        if re.search(r'(pathetic|worthless|disgusting|trash|garbage|loser|scum|evil|vile)', tl):
            cats.add('Dehumanizing')
        if re.search(r'(insane|crazy|lunatic|psycho|unhinged|deranged|mental|nuts)', tl):
            if not is_colloquial_crazy:
                cats.add('Calling her crazy/insane')
        if re.search(r'(grift|con\s*artist|showboat|performative|virtue\s*signal)', tl):
            cats.add('Grifting accusation')
        if re.search(r'(incompetent|useless|feckless|do\s*nothing|does\s*nothing|lazy)', tl):
            cats.add('Competence attack')
        if re.search(r'(shut\s*up|nobody\s*cares|no\s*one\s*cares|sit\s*down)', tl):
            cats.add('Aggressive dismissal')
        if re.search(r'(deport|illegal.*immigrant|ice\b)', tl):
            cats.add('Racist comment')

    # Second fallback
    if not cats:
        if re.search(r'(bad|terrible|awful|wrong|horrible|shitty|weak|coward)', tl):
            if re.search(r'(you|she|her|jackie|fielder)', tl):
                cats.add('Exaggerated criticism')
        elif re.search(r'(lol|lmao|haha|😂|🤣|🤡)', tl):
            cats.add('Condescension')
        elif re.search(r'(not\s*serious|credibility|joke)', tl):
            cats.add('Condescension')
        else:
            cats.add('Exaggerated criticism')

    return '; '.join(sorted(cats))


def rate_hostility(text):
    """
    Score tweet hostility on a 1-10 scale.
    Additive scoring with category-based floors.
    """
    tl = text.lower()
    score = 0
    is_caps = (
        text.isupper() or
        (len(text) > 10 and
         sum(1 for c in text if c.isupper()) / max(len(text.replace(' ', '')), 1) > 0.6)
    )

    # --- Tier 1: 5 pts — death/harm wishes ---
    if re.search(r'(should|hope|wish|pray|needs?\s*to|deserves?\s*to)\s*(die|drop\s*dead|burn|rot|suffer|be\s*(killed|shot|hanged|executed))', tl): score += 5
    if re.search(r'kill\s*(yourself|herself)|euthanasia|put\s*(her|you)\s*down', tl): score += 5
    if re.search(r'suffers?\s*miserably', tl): score += 5
    if re.search(r'(send|release).*(gang|criminal|dealer).*your\s*(house|home)|earthquake.*fix\s*everything', tl): score += 5
    if re.search(r'burn\s*in\s*hell', tl): score += 5
    if re.search(r'(shoot|shooting)\s*(california\s*politic|politic)', tl): score += 5

    # --- Tier 2: 5 pts — slurs ---
    if re.search(r'\bcunt\b|\bcvnt\b|\bcuntbitch\b', tl): score += 5
    if re.search(r'\bwhore\b|\bslut\b|\bskank\b|\bputa\b|\bsalesputa\b', tl): score += 5
    if re.search(r'\bbitch\b|\bbiatch\b|\bbiotch\b|\bbitches\b', tl): score += 5
    if re.search(r'\btwat\b|\bho\s*bag\b|\bhag\b', tl): score += 5
    if re.search(r'\bwench\b', tl): score += 6  # weighted higher per Sasha
    if re.search(r'\bhoe\b', tl): score += 5
    if re.search(r'\bretard(ed|s|io|ios)?\b|@ifindretards|re\.\s*tard|full\s*retard', tl): score += 5
    if re.search(r'\bdyke\b|\blesbo\b|\bbutch\b', tl): score += 5
    if re.search(r'wild\s*in\s*the\s*sack|bikini.*lingerie|put.*in.*bikini|sugar\s*daddy.*hmu|hot\s*teacher.*hmu|pack\s*lightly.*bikini', tl): score += 5
    if re.search(r'easy\s*on\s*the\s*eyes|she.s\s*(hot|cute|attractive)|hotttt|hot\s*chick|eye\s*candy|🥵', tl): score += 5

    # --- Tier 3: 4 pts — dehumanizing, appearance ---
    if re.search(r'(you.re|she.s|you\s*are|she\s*is|jackie\s*is)\s*(worthless|trash|garbage|disgusting|evil|vile|scum|despicable|pathetic|a\s*bad\s*person|a\s*cancer|a\s*stain|a\s*disgrace)', tl): score += 4
    if re.search(r'worthless\s*(jackie|fielder|supervisor|politician|elected)', tl): score += 4
    if re.search(r'(\btrash\b|\bgarbage\b|\bscum\b|\bghoul\b|\bparasite\b|\bleech\b)', tl) and re.search(r'(you|she|her|jackie|fielder)', tl): score += 4
    if re.search(r'waste\s*of\s*(space|skin|oxygen|air|time|water|calcium)', tl): score += 4
    if re.search(r'piece\s*of\s*(shit|crap|garbage)|barely\s*human|subhuman|sludge', tl): score += 4
    if re.search(r'(evil|cancer|tumor|plague|poison|stain)\b', tl) and re.search(r'(jackie|fielder|you|she|supervisor|your\s*bio)', tl): score += 4
    if re.search(r'(smell|stink|smelly|stank|shower|arm\s*pit)', tl) and re.search(r'(you|she|her|screen|through|jackie|fielder)', tl): score += 4
    if re.search(r'quiet\s*piggy|\boink\b|piggy', tl): score += 4
    if re.search(r'\bugly\b|skeletor|five\s*head|forehead.*(large|alarming)|got\s*ugly', tl): score += 4
    if re.search(r'mentally\s*(ill|unstable)\s*(woman|female)', tl): score += 4
    if re.search(r'\bso\s*gay\b|Gay!\b', tl): score += 4
    if re.search(r'fake\s*indigenous|pretendian|fake\s*indian', tl): score += 4
    if re.search(r'testicular\s*cancer', tl): score += 3

    # --- Tier 4: 3 pts — profanity, extreme criticism ---
    if re.search(r'\bfuck\b|\bfucking\b|\bfuckin\b|\bfking\b|\bf\*ck|fekk|fck\b', tl): score += 3
    if re.search(r'fuck\s*(you|off|yourself|her)\b|go\s*fuck|🖕|gfy|stfu|gtfo|fu\s+loser|\bfu\b$', tl): score += 1
    if re.search(r'(destroying|ruining|crushed|killed)\s*(sf|san\s*fran|the\s*(city|district|state)|california)', tl): score += 3
    if re.search(r'blood\s*on\s*(your|her)\s*hands', tl): score += 3
    if re.search(r'committed\s*great\s*evils', tl): score += 3

    # --- Tier 5: 2 pts ---
    if re.search(r'\b(idiot|moron|imbecile|halfwit|dimwit|nitwit|dunce|dipshit|dipstick|dumbass|jackass|jackoff|pea.brain|tool\s*fool|midwit|maroon|bozo)\b', tl): score += 2
    if re.search(r'\bstupid\b|\bdumb\b|\bdumbest\b|\bdopey\b|\bdimb\b|brain\s*dead|economically\s*illiterate|financially\s*illiterate', tl): score += 2
    if re.search(r'🤡', tl): score += 2
    if re.search(r'\bloser\b', tl): score += 2
    if re.search(r'\b(insane|crazy|lunatic|psycho|unhinged|deranged|delusional|loony|bonkers|wacko|nutjob|nutter|batshit)\b', tl): score += 2
    if re.search(r'mentally\s*(ill|unstable|unfit|unbalanced)', tl): score += 2
    if re.search(r'shut\s*(the\s*fuck\s*)?up|zip\s*it|stfu', tl): score += 2
    if re.search(r'nobody\s*(likes|cares|wants)|no\s*one\s*(cares|takes)', tl): score += 2
    if re.search(r'(should|needs?\s*to)\s*(be\s*)?(in\s*jail|arrested|prosecuted)', tl): score += 2
    if re.search(r'broke\s*the\s*law|breaking\s*the\s*law|illegally\s*leak', tl): score += 2
    if re.search(r'deport\s*(this|her|you|jackie|fielder)|leave\s*america|get\s*out\s*of\s*(my|our)\s*country|move\s*to\s*(cuba|venezuela|minnesota)', tl): score += 2
    if re.search(r'(worst|most\s*(useless|incompetent))\s*(super|politi|person|human)', tl): score += 2
    if re.search(r'(she|you|jackie|fielder)\s*(are|is)\s*(the|a|part\s*of\s*the)\s*problem', tl): score += 2
    if re.search(r'childless|empty\s*egg\s*carton|abolish\s*the\s*19th|women\s*don.t\s*belong|big\s*girl\s*job', tl): score += 2
    if re.search(r'(faking|fake|staged|stunt).{0,40}(hospital|mental|illness|crisis)', tl): score += 2
    if re.search(r'god\s*is\s*good|lord\s*works|good\s*riddance|cause\s*for\s*celebration|bye\s*felecia|nature\s*is\s*healing', tl): score += 2
    if re.search(r'#?hitlerbarbie|banshee|socialista|princess\s*of\s*death|girl\s*boss|hysterical', tl): score += 2
    if re.search(r'(commie|communist|marxist)\s*(trash|scum|filth|rat|pig|bastard|hippie|piece)', tl): score += 2
    if re.search(r'bipolar|schizophren|borderline|cutter|britney\s*spears|adachi|narcissis|psychosis|looks?\s*high', tl): score += 2
    if re.search(r'temper\s*tantrum|(?:she|jackie)\s*is\s*a\s*child|poor\s*baby|overgrown\s*child|adulting|theater\s*kid', tl): score += 2
    if re.search(r'\bestupida\b|\bpendeja\b|\bputa\b|\bsalesputa\b', tl): score += 2
    if re.search(r'crook\b|crooked\b', tl): score += 2
    if re.search(r'disgrace\s*to\s*(the|human)', tl): score += 2
    if re.search(r'(i\s*|we\s*)(hate|loathe|despise|detest)\s*(you|her|jackie|fielder)', tl): score += 2
    if re.search(r'pap\s*smear|cumunist', tl): score += 2
    if re.search(r'drooling', tl): score += 2

    # --- Tier 6: 1 pt ---
    if re.search(r'(incompetent|useless|feckless|clueless|inept|do.nothing|negligent)', tl): score += 1
    if re.search(r'\bgrift(er|ing|s)?\b|showboat|performative|champagne\s*socialist|virtue\s*signal', tl): score += 1
    if re.search(r'(corrupt|criminal|fraud|kleptocracy)\b', tl): score += 1
    if re.search(r'embarrassment|embarrass', tl): score += 1
    if re.search(r'(recall|fire|remove|oust|vote\s*out|get\s*rid|kick\s*out|resign|step\s*down)', tl): score += 1
    if re.search(r'(commie|communist|marxist|bolshevik)\b', tl) and not re.search(r'(commie|communist|marxist)\s*(trash|scum|filth|rat|pig|bastard|cunt|bitch)', tl): score += 1
    if re.search(r'(nazi|antisemit|anti.semit|jew.hat|anti.jewish|ham\s*ass)', tl): score += 1
    if re.search(r'(does?\s*nothing|jack\s*shit|what.*do\s*all\s*day|do\s*your\s*job|do\s*something|do\s*better|absent|no\s*show)', tl): score += 1
    if re.search(r'(liar|lying|full\s*of\s*(shit|it)|dishonest|gaslighting|bullshit|stop\s*lying)', tl): score += 1
    if re.search(r'(carpetbag|white\s*divorc|te\s*falta|stanford\s*fresa)', tl): score += 1
    if re.search(r'(tds\b|waymo\s*derangement|meds\s*kick|mental\s*ward|psych\s*unit|she\s*cray)', tl): score += 1
    if re.search(r'(socialist|leftist|liberal|dsa|democrat).{0,20}(insane|crazy|mental|lunatic|nuts|unstable)', tl): score += 1
    if re.search(r'coward|miserable|not\s*serious|joke\s*of\s*a', tl): score += 1
    if re.search(r'(district|mission|city|sf)\s*(is\s*)?(a\s*)?(mess|disaster|filthy|dirty|unsafe|failing|rotten|shithole|dumpster|sewer|toilet|hell)', tl): score += 1
    if re.search(r'(condescen|patroniz|🙄)', tl): score += 1
    if re.search(r'(take|try)\s*(a|some)\s*econ|read\s*(the\s*)?constitution|welcome\s*to\s*socialism|tell\s*me\s*you.*without|seek\s*help', tl): score += 1
    if re.search(r'(not\s*(qualified|fit|competent|capable)|zero\s*experience|no\s*(clue|idea|concept))', tl): score += 1
    if re.search(r'(shit|crap|shitty|shithead|shithole|ass|asshole|azwipe)\b', tl): score += 1
    if re.search(r'(bad|terrible|awful|horrible|worst)\s*(person|human|supervisor|politician)', tl): score += 1
    if re.search(r'\bkarma\b', tl): score += 1
    if re.search(r'(cringe|corny|laughable|ridiculous|absurd)', tl): score += 1
    if re.search(r'(why\s*we\s*(have|got))\s*(trump|maga)', tl): score += 1
    if re.search(r'(deport|illegal|ice\b|undocumented|border)', tl): score += 1
    if re.search(r'job\s*killer', tl): score += 1
    if re.search(r'wtf\b|azwipe', tl): score += 1
    if re.search(r'demon\b', tl): score += 1
    if re.search(r'not\s*allowed\s*to\s*vote', tl): score += 1

    # ALL CAPS bonus
    if is_caps and len(text) > 10:
        score += 1

    # --- Floors (minimum scores for certain categories) ---
    # Dehumanizing = min 5
    if re.search(r'(you.re|she.s|you\s*are|she\s*is)\s*(worthless|trash|garbage|evil|disgusting|vile|scum|pathetic|a\s*bad\s*person|a\s*cancer)', tl): score = max(score, 5)
    if re.search(r'worthless\s*(jackie|fielder|supervisor)', tl): score = max(score, 5)
    if re.search(r'(\btrash\b|\bgarbage\b|\bscum\b|\bghoul\b|\bparasite\b|\bleech\b)', tl) and re.search(r'(you|she|her|jackie|fielder)', tl): score = max(score, 5)
    if re.search(r'waste\s*of\s*(space|skin|oxygen|air|time|water|calcium)', tl): score = max(score, 5)
    if re.search(r'piece\s*of\s*(shit|crap|garbage)|barely\s*human|subhuman|sludge', tl): score = max(score, 5)
    # Appearance/hygiene = min 5
    if re.search(r'(smell|stink|smelly|stank|shower|arm\s*pit)', tl) and re.search(r'(you|she|her|screen|through|jackie|fielder)', tl): score = max(score, 5)
    if re.search(r'quiet\s*piggy|\boink\b|piggy', tl): score = max(score, 5)
    if re.search(r'\bugly\b|skeletor|five\s*head|forehead.*(large|alarming)|got\s*ugly', tl): score = max(score, 5)
    # Slurs = min 5
    if re.search(r'\bcunt\b|\bcvnt\b|\bbitch\b|\bbiatch\b|\bwhore\b|\bslut\b|\bskank\b|\btwat\b|\bhoe\b|\bwench\b|\bhag\b|\bputa\b|\bsalesputa\b', tl): score = max(score, 5)
    if re.search(r'\bretard(ed|s|io|ios)?\b|@ifindretards|full\s*retard', tl): score = max(score, 5)
    if re.search(r'\bdyke\b|\blesbo\b|\bbutch\b', tl): score = max(score, 5)
    # Insult + political = min 5
    if re.search(r'\b(halfwit|dimwit|nitwit|idiot|moron|imbecile|dumbass|dipshit)\b', tl):
        if re.search(r'(socialist|communist|commie|leftist|liberal|marxist|has\s*thoughts)', tl):
            score = max(score, 5)
    # Death/harm = min 6
    if re.search(r'(should|hope|wish|needs?\s*to|deserves?\s*to)\s*(die|drop\s*dead|burn|rot|suffer|be\s*(killed|shot|hanged))|suffers?\s*miserably|euthanasia|burn\s*in\s*hell', tl): score = max(score, 6)
    if re.search(r'(shoot|shooting)\s*(california\s*politic|politic)', tl): score = max(score, 6)
    # Sexual harassment = min 5
    if re.search(r'sugar\s*daddy.*hmu|hot\s*teacher.*hmu|put.*in.*bikini|pack\s*lightly.*bikini', tl): score = max(score, 5)

    if score == 0:
        score = 1
    return min(score, 10)


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_xlsx(df, output_path):
    """Write categorized DataFrame to formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Categorized Tweets"

    col_order = [
        'Display Name', 'Username', 'Date', 'Reply Text', 'Likes',
        'Flag', 'Category', 'Hostility (1-10)', 'Reply Type', 'Notes',
    ]
    field_map = {
        'Display Name': 'Display Name', 'Username': 'Username',
        'Date': 'Date', 'Reply Text': 'Reply Text',
        'Likes': 'Likes', 'Flag': 'Flag', 'Category': 'Category',
        'Hostility (1-10)': 'Hostility (1-10)',
        'Reply Type': 'Reply Type', 'Notes': 'Notes',
    }

    header_fill = PatternFill('solid', fgColor='2B5797')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    data_font = Font(name='Arial', size=11)
    alt_fill = PatternFill('solid', fgColor='F2F2F2')

    # Headers
    for col_idx, h in enumerate(col_order, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Data rows
    row_num = 2
    for _, r in df.iterrows():
        for col_idx, col_name in enumerate(col_order, 1):
            val = r.get(col_name, '')
            if pd.isna(val):
                val = ''
            if col_name == 'Likes':
                try:
                    val = int(val)
                except (ValueError, TypeError):
                    val = 0
            if col_name == 'Reply Text':
                val = str(val)[:2000]

            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical='center',
                wrap_text=(col_idx in (4, 7, 10)),
            )
            if row_num % 2 == 0 and col_idx not in (6, 8):
                cell.fill = alt_fill

        # Flag coloring
        flag = str(r.get('Flag', '')).strip().lower()
        if flag == 'yes':
            ws.cell(row=row_num, column=6).fill = PatternFill('solid', fgColor='FCE4EC')
        elif flag == 'no':
            ws.cell(row=row_num, column=6).fill = PatternFill('solid', fgColor='E8F5E9')
        elif flag in ('h', 'maybe'):
            ws.cell(row=row_num, column=6).fill = PatternFill('solid', fgColor='FFF3CD')

        # Hostility coloring
        h = r.get('Hostility (1-10)', '')
        if h != '' and h is not None and str(h).strip() not in ('', 'nan'):
            try:
                hv = int(float(h))
                hc = ws.cell(row=row_num, column=8)
                if hv >= 8:
                    hc.fill = PatternFill('solid', fgColor='C0392B')
                    hc.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                elif hv >= 6:
                    hc.fill = PatternFill('solid', fgColor='E74C3C')
                    hc.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
                elif hv >= 4:
                    hc.fill = PatternFill('solid', fgColor='F39C12')
                    hc.font = Font(name='Arial', size=11, bold=True)
                elif hv >= 2:
                    hc.fill = PatternFill('solid', fgColor='F7DC6F')
                    hc.font = Font(name='Arial', size=11, bold=True)
                else:
                    hc.fill = PatternFill('solid', fgColor='FCF3CF')
                    hc.font = Font(name='Arial', size=11, bold=True)
                hc.alignment = Alignment(horizontal='center', vertical='center')
            except (ValueError, TypeError):
                pass

        row_num += 1

    # Column widths
    widths = {'A': 26, 'B': 22, 'C': 14, 'D': 65, 'E': 8,
              'F': 10, 'G': 55, 'H': 16, 'I': 16, 'J': 45}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:J{row_num - 1}"
    wb.save(output_path)
    return row_num - 2


# ============================================================
# MAIN
# ============================================================

def main():
    if len(sys.argv) < 3:
        print("Usage: python categorize_tweets.py input.csv output.xlsx")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    print(f"Reading {input_path}...")
    df = pd.read_csv(input_path)
    df['Flag'] = df['Flag'].fillna('')

    # Ensure all columns exist
    for col in ['Display Name', 'Username', 'Date', 'Reply Text', 'Likes',
                'Flag', 'Category', 'Hostility (1-10)', 'Reply Type', 'Notes']:
        if col not in df.columns:
            df[col] = ''

    yes_count = 0
    maybe_count = 0
    skip_count = 0

    for idx, row in df.iterrows():
        flag = str(row['Flag']).strip().lower()
        text = str(row['Reply Text']) if pd.notna(row['Reply Text']) else ''

        if flag in ('yes',):
            # Categorize and score
            df.at[idx, 'Category'] = categorize(text)
            df.at[idx, 'Hostility (1-10)'] = rate_hostility(text)
            yes_count += 1

        elif flag in ('no',):
            # Skip
            skip_count += 1

        else:
            # Test against patterns; if any match, flag as Maybe
            test_cat = categorize(text)
            if test_cat and test_cat != 'Exaggerated criticism':
                df.at[idx, 'Flag'] = 'Maybe'
                df.at[idx, 'Category'] = test_cat
                df.at[idx, 'Hostility (1-10)'] = rate_hostility(text)
                maybe_count += 1

    print(f"\nProcessed {len(df)} tweets:")
    print(f"  Yes (categorized): {yes_count}")
    print(f"  Maybe (new): {maybe_count}")
    print(f"  No (skipped): {skip_count}")
    print(f"  Other: {len(df) - yes_count - maybe_count - skip_count}")

    # Category stats
    all_cats = []
    yes_df = df[df['Flag'].str.lower().isin(['yes'])]
    for _, row in yes_df.iterrows():
        cat = str(row['Category']) if pd.notna(row['Category']) else ''
        for c in cat.split(';'):
            c = c.strip()
            if c:
                all_cats.append(c)

    print(f"\n=== CATEGORY BREAKDOWN ({len(set(all_cats))} categories) ===")
    for cat, count in Counter(all_cats).most_common():
        print(f"  {cat}: {count}")

    # Write output
    print(f"\nWriting {output_path}...")
    rows_written = write_xlsx(df, output_path)
    print(f"Done. {rows_written} rows written.")


if __name__ == '__main__':
    main()